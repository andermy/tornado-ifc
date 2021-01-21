import pandas as pd
from openpyxl import load_workbook
import mongo as m
from bson.objectid import ObjectId

class Excel():
    def __init__(self, file=None, column_row=None):
        self.path = file['path']
        self.sheet_names = self.get_sheet_names()
        self.df = []
        self.project = file['version']['project']

    def get_sheet_names(self):
        wb = load_workbook(filename = self.path)
        return wb.sheetnames
    
    def get_sheets(self, column_row=None):
        for sheet in self.sheet_names:
            df = pd.read_excel(self.path, header=column_row, sheet_name=sheet)
            self.df.append(Table(df, sheet, self.project))

class Table():
    def __init__(self, dataframe=None, sheet_name=None, version):
        self.sheet = sheet_name
        self.df = dataframe
        self.version = version
        self.db = m.MongoDb()
    
    def create_from_ifc(self, db_type):
        self.db.define_collection(db_type)
        db_type_doc = self.db.query({'version': ObjectId(self.version)})
        new_doc = {}
        for item in db_type_doc[0].keys():
            new_doc[item] = list(map(operator.itemgetter(item), db_type_doc))
        
        return new_doc

    def create_in_db(self):
        self.db.define_collection('frame')
        to_dict = self.dataframe.to_dict()
        self.db.insert_one(to_dict)

    def update_in_db(self):
        self.db.define_collection('frame')
        to_dict = self.dataframe.to_dict()
        self.db.update_one(to_dict)

    def calculation(self, calculation):
        try:
            self.dataframe[calculation['created_variable']] = eval(calculation['recipe'])
            self.update_in_db()
        except:
            pass